import openpyxl
from openpyxl.styles import Font, PatternFill

# Sample file 1
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = "示例数据"
ws1.append(["编号", "名称", "数量", "日期"])
ws1.append([1, "春茶", 120, "2026-03-15"])
ws1.append([2, "夏茶", 85, "2026-06-20"])
ws1.append([3, "秋茶", 95, "2026-09-10"])
header_font = Font(bold=True)
header_fill = PatternFill(start_color="E7E6E1", end_color="E7E6E1", fill_type="solid")
for cell in ws1[1]:
    cell.font = header_font
    cell.fill = header_fill
wb1.save(r"C:\Users\13521\fox-spring-tea-website\files\sample-data-01.xlsx")

# Sample file 2
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "统计汇总"
ws2.append(["类别", "销售额", "成本", "利润"])
ws2.append(["绿茶", 15800, 9200, 6600])
ws2.append(["红茶", 23100, 12800, 10300])
ws2.append(["乌龙茶", 18750, 10500, 8250])
for cell in ws2[1]:
    cell.font = header_font
    cell.fill = header_fill
wb2.save(r"C:\Users\13521\fox-spring-tea-website\files\sample-data-02.xlsx")

print("Excel files created")
