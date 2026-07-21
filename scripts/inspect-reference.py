import openpyxl
import json

wb = openpyxl.load_workbook(
    r"C:\Users\13521\Desktop\随想录\1-Projects（有终点的事情）\实习\test_data.xlsx",
    data_only=True,
)
ws = wb.active

headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
rows = []
for r in range(2, ws.max_row + 1):
    row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    rows.append(row)

result = {
    "sheet": ws.title,
    "max_row": ws.max_row,
    "max_column": ws.max_column,
    "headers": headers,
    "rows": rows,
}

output_path = r"C:\Users\13521\fox-spring-tea-website\scripts\reference-structure.json"
with open(output_path, "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2, default=str)

print(f"Saved structure to {output_path}")
