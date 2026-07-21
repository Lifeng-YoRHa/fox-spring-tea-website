import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta

def generate_sales_data(start_date, skus, platforms, base_visits, base_orders, base_refund, unit_prices):
    """Generate ~30 rows of sales data matching the reference format."""
    rows = []
    dates = [start_date + timedelta(days=i) for i in range(5)]
    for day_index, date in enumerate(dates):
        visits = base_visits + day_index * 10
        orders = base_orders + day_index
        refund = base_refund + day_index * 2.5
        for sku in skus:
            for platform in platforms:
                sales = round(orders * unit_prices[sku], 2)
                rows.append([
                    sku,
                    platform,
                    date.strftime("%Y-%m-%d"),
                    visits,
                    orders,
                    sales,
                    refund,
                ])
    return rows


def write_sales_workbook(path, sheet_name, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header
    ws.append(headers)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E7E6E1", end_color="E7E6E1", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for row in rows:
        ws.append(row)

    # Auto-adjust column widths roughly
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 4, 20)

    wb.save(path)


headers = ["商品编码", "平台", "日期", "访问量", "订单量", "销售额", "退款金额"]
platforms = ["天猫", "抖音"]

# File 1: July data
skus_1 = ["SKU001", "SKU002", "SKU003"]
unit_prices_1 = {"SKU001": 170.1, "SKU002": 116.1, "SKU003": 2520.0}
rows_1 = generate_sales_data(
    datetime(2026, 7, 1), skus_1, platforms,
    base_visits=100, base_orders=5, base_refund=12.5,
    unit_prices=unit_prices_1,
)
write_sales_workbook(
    r"C:\Users\13521\fox-spring-tea-website\files\sample-data-01.xlsx",
    "销售数据",
    headers,
    rows_1,
)

# File 2: August data with different SKUs
skus_2 = ["SKU004", "SKU005", "SKU006"]
unit_prices_2 = {"SKU004": 210.5, "SKU005": 145.8, "SKU006": 3150.0}
rows_2 = generate_sales_data(
    datetime(2026, 8, 1), skus_2, platforms,
    base_visits=120, base_orders=6, base_refund=15.0,
    unit_prices=unit_prices_2,
)
write_sales_workbook(
    r"C:\Users\13521\fox-spring-tea-website\files\sample-data-02.xlsx",
    "销售数据",
    headers,
    rows_2,
)

print(f"sample-data-01.xlsx: {len(rows_1)} data rows")
print(f"sample-data-02.xlsx: {len(rows_2)} data rows")
print("Done")
